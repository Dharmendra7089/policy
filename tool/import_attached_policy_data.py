import argparse
import json
import math
import os
import re
import subprocess
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from urllib import error, request

from openpyxl import load_workbook


PROJECT_ID = "insurance-1178e"
BASE_URL = (
    f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}"
    "/databases/(default)/documents"
)

FILES = {
    "General": Path(r"C:\Users\user\Downloads\General Insurance Policies Data.xlsx"),
    "Life": Path(r"C:\Users\user\Downloads\Life Insurance Policies Data.xlsx"),
    "Health": Path(r"C:\Users\user\Downloads\Health Insurance Policies Data.xlsx"),
}

COLLECTIONS_TO_REPLACE = [
    "insurance_departments",
    "insurance_companies",
    "policies",
    "life_policies",
    "general_policies",
    "customers",
    "customer_policies",
    "revenue",
    "lead_serial_registry",
]


COMPANY_ALIASES = {
    "IFFCO TOKIO": "IFFCO Tokio General Insurance Company Limited",
    "SBI GENERAL": "SBI General Insurance Company Limited",
    "SBI MOTOR": "SBI General Insurance Company Limited",
    "BAJAJ GENERAL": "Bajaj Allianz General Insurance Company Limited",
    "HDFC ERGO": "HDFC ERGO General Insurance Company Limited",
    "ICICI LOMBARD": "ICICI Lombard General Insurance Company Limited",
    "UNITED INDIA": "United India Insurance Company Limited",
    "NEW INDIA": "The New India Assurance Company Limited",
    "BAJAJ LIFE": "Bajaj Allianz Life Insurance Company Limited",
    "BAJAJ ALLIANZ LIFE": "Bajaj Allianz Life Insurance Company Limited",
    "ICICI PRU": "ICICI Prudential Life Insurance Company Limited",
    "ICICI PRUDENTIAL": "ICICI Prudential Life Insurance Company Limited",
    "ADITYA BIRLA SUN LIFE": "Aditya Birla Sun Life Insurance Company Limited",
    "AXIS MAX": "Axis Max Life Insurance Limited",
    "MAX LIFE": "Axis Max Life Insurance Limited",
    "SBI LIFE": "SBI Life Insurance Company Limited",
    "INDUSIND NIPPON": "IndusInd Nippon Life Insurance Company Limited",
    "INDUS IND NIPPON": "IndusInd Nippon Life Insurance Company Limited",
    "NIVA BUPA": "Niva Bupa Health Insurance Company Limited",
    "ABHI": "Aditya Birla Health Insurance Company Limited",
    "ADITYA BIRLA HEALTH": "Aditya Birla Health Insurance Company Limited",
    "CARE HEALTH": "Care Health Insurance Limited",
    "STAR HEALTH": "Star Health and Allied Insurance Company Limited",
    "GALAXY HEALTH": "Galaxy Health Insurance Company Limited",
    "MANIPAL CIGNA": "ManipalCigna Health Insurance Company Limited",
    "MANIPALCIGNA": "ManipalCigna Health Insurance Company Limited",
}


def slug(value):
    text = str(value or "").lower().replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "item"


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return re.sub(r"\s+", " ", text)


def number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return 0.0
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", str(value))
    if text in {"", "-", "."}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)) and not math.isnan(value):
        # Excel serial date.
        return datetime(1899, 12, 30) + timedelta(days=int(value))
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    short = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2})", text)
    if short:
        day, month, year = map(int, short.groups())
        year += 2000
        return datetime(year, month, day)
    return None


def parse_month_start(value):
    text = clean_text(value)
    if not text:
        return None
    if isinstance(value, datetime):
        return datetime(value.year, value.month, 1)
    match = re.search(
        r"(jan|feb|mar|apr|april|may|jun|june|jul|july|aug|sep|sept|oct|nov|dec)[a-z]*'?[-\s]*(\d{2,4})",
        text.lower(),
    )
    if not match:
        return None
    month_names = {
        "jan": 1,
        "feb": 2,
        "mar": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "sep": 9,
        "sept": 9,
        "oct": 10,
        "nov": 11,
        "dec": 12,
    }
    year = int(match.group(2))
    if year < 100:
        year += 2000
    key = match.group(1)
    return datetime(year, month_names[key], 1)


def one_year_minus_one_day(start):
    try:
        return datetime(start.year + 1, start.month, start.day) - timedelta(days=1)
    except ValueError:
        return datetime(start.year + 1, start.month, 28) - timedelta(days=1)


def iso(dt):
    if not dt:
        return None
    return dt.strftime("%Y-%m-%dT00:00:00Z")


def fmt_date(dt):
    if not dt:
        return ""
    return dt.strftime("%d.%m.%Y")


def month_key(dt):
    if not dt:
        return ""
    return f"{dt.year}-{dt.month:02d}"


def infer_company(sheet_name, workbook_category, rows):
    search = " ".join(clean_text(cell).upper() for row in rows[:6] for cell in row)
    search = f"{sheet_name.upper()} {search}"
    for key, canonical in COMPANY_ALIASES.items():
        if key in search:
            return canonical
    return clean_text(sheet_name)


def infer_category(workbook_category, sheet_name, company_name, product):
    text = f"{sheet_name} {company_name} {product}".upper()
    if workbook_category == "Life":
        return "Life"
    if workbook_category == "Health":
        return "Health"
    if "HEALTH" in text or "MEDICLAIM" in text or "GMC" in text or "GPA" in text:
        return "Health"
    return "General"


def departments_for_company(company, categories):
    if company == "SBI General Insurance Company Limited":
        return ["General", "Motor", "Health"]
    if any(cat == "Life" for cat in categories):
        return ["Life"]
    depts = set(categories)
    if company in {
        "IFFCO Tokio General Insurance Company Limited",
        "Bajaj Allianz General Insurance Company Limited",
        "HDFC ERGO General Insurance Company Limited",
        "ICICI Lombard General Insurance Company Limited",
        "United India Insurance Company Limited",
        "The New India Assurance Company Limited",
    }:
        depts.add("General")
    return sorted(depts)


def find_header(ws):
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        values = [clean_text(ws.cell(row_idx, col).value).lower() for col in range(1, ws.max_column + 1)]
        joined = " ".join(values)
        if "sl no" in joined and "policy" in joined:
            return row_idx, values
    return None, []


def header_map(headers):
    result = {}
    for index, label in enumerate(headers):
        label = label.lower()
        if "sl no" in label:
            result["sl_no"] = index
        elif "name" in label and "person" in label:
            result["customer_name"] = index
        elif "policy no" in label or label == "policy":
            result["policy_number"] = index
        elif label == "month" or "month" in label:
            result["month"] = index
        elif "issued" in label or "issue" in label:
            result["issue_date"] = index
        elif "end date" in label or "policy end" in label or "expiry" in label:
            result["end_date"] = index
        elif "sum assured" in label:
            result["sum_assured"] = index
        elif "premium" in label:
            result["premium"] = index
        elif "phone" in label or "mobile" in label:
            result["phone"] = index
        elif "product" in label:
            result["product"] = index
        elif "base commission" in label:
            result["commission_percent"] = index
        elif label == "commission" or "commission" in label:
            result.setdefault("commission_amount", index)
        elif "gst" in label or "cgst" in label or "sgst" in label:
            result.setdefault("gst", index)
        elif "total" in label or "net total" in label:
            result.setdefault("net_total", index)
    return result


def parse_workbooks():
    rows_out = []
    skipped = []
    sheet_masters = []
    for workbook_category, path in FILES.items():
        if not path.exists():
            raise FileNotFoundError(path)
        wb = load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            raw_rows = [
                [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, min(ws.max_row, 8) + 1)
            ]
            company = infer_company(ws.title, workbook_category, raw_rows)
            sheet_category = infer_category(workbook_category, ws.title, company, ws.title)
            sheet_masters.append(
                {
                    "sourceWorkbook": path.name,
                    "sourceSheet": ws.title,
                    "companyName": company,
                    "category": sheet_category,
                    "product": f"{ws.title} Portfolio",
                }
            )
            header_idx, headers = find_header(ws)
            if not header_idx:
                skipped.append({"sheet": ws.title, "reason": "header not found"})
                continue
            hm = header_map(headers)
            required = {"sl_no", "customer_name", "policy_number"}
            if not required.issubset(hm):
                skipped.append({"sheet": ws.title, "reason": f"missing {sorted(required - set(hm))}"})
                continue
            for r in range(header_idx + 1, ws.max_row + 1):
                values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                sl = clean_text(values[hm["sl_no"]]) if hm.get("sl_no") is not None else ""
                if not re.fullmatch(r"\d+", sl):
                    continue
                name = clean_text(values[hm["customer_name"]])
                policy_no = clean_text(values[hm["policy_number"]]).upper()
                if not name or not policy_no:
                    continue
                issue = parse_date(values[hm["issue_date"]]) if "issue_date" in hm else None
                end = parse_date(values[hm["end_date"]]) if "end_date" in hm else None
                if not issue and not end and "month" in hm:
                    issue = parse_month_start(values[hm["month"]])
                    if issue:
                        end = datetime(issue.year + 1, issue.month, 1) - timedelta(days=1)
                if not issue and end:
                    issue = datetime(end.year - 1, end.month, min(end.day + 1, 28))
                if issue and not end:
                    end = one_year_minus_one_day(issue)
                if issue and end and end <= issue:
                    end = one_year_minus_one_day(issue)
                product = clean_text(values[hm["product"]]) if "product" in hm else ""
                if not product or re.fullmatch(r"\d{10}", product):
                    product = f"{ws.title} Portfolio"
                category = infer_category(workbook_category, ws.title, company, product)
                rows_out.append(
                    {
                        "sourceWorkbook": path.name,
                        "sourceSheet": ws.title,
                        "sourceRow": r,
                        "companyName": company,
                        "category": category,
                        "customerName": name,
                        "policyNumber": policy_no,
                        "month": clean_text(values[hm["month"]]) if "month" in hm else month_key(issue),
                        "issueDate": issue,
                        "policyEndDate": end,
                        "sumAssured": number(values[hm["sum_assured"]]) if "sum_assured" in hm else 0,
                        "premium": number(values[hm["premium"]]) if "premium" in hm else 0,
                        "phone": clean_text(values[hm["phone"]]) if "phone" in hm else "",
                        "product": product,
                        "commissionPercent": number(values[hm["commission_percent"]]) * 100
                        if 0 < number(values[hm.get("commission_percent", -1)] if "commission_percent" in hm else 0) <= 1
                        else number(values[hm["commission_percent"]]) if "commission_percent" in hm else 0,
                        "commissionAmount": number(values[hm["commission_amount"]]) if "commission_amount" in hm else 0,
                        "gst": number(values[hm["gst"]]) if "gst" in hm else 0,
                        "netTotal": number(values[hm["net_total"]]) if "net_total" in hm else 0,
                    }
                )
    return rows_out, skipped, sheet_masters


def company_invoice_code(company):
    parts = re.findall(r"[A-Za-z0-9]+", company.upper())
    skip = {"INSURANCE", "COMPANY", "LIMITED", "GENERAL", "LIFE", "HEALTH", "AND", "THE"}
    useful = [p for p in parts if p not in skip]
    code = "".join(p[:3] if len(useful) == 1 else p[0] for p in useful[:3])
    return f"MAKK_{code or 'INV'}"


def category_code(category):
    key = clean_text(category).lower()
    if key == "health":
        return "H"
    if key == "life":
        return "L"
    if key == "general":
        return "G"
    if key in {"agriculture", "agricultural"}:
        return "A"
    if key == "ecgc":
        return "E"
    return "G"


def lead_fields(category, year, sequence):
    code = category_code(category)
    lead_id = f"M{code}{year}{sequence:05d}"
    return lead_id, {
        "leadUniqueId": lead_id,
        "uniqueLeadId": lead_id,
        "leadSerialNumber": lead_id,
        "leadSerialSequence": sequence,
        "leadSerialYear": year,
        "leadSerialCategoryCode": code,
    }


def build_documents(rows, sheet_masters):
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    serial_year = 2026
    company_categories = {}
    for sheet in sheet_masters:
        company_categories.setdefault(sheet["companyName"], set()).add(sheet["category"])
    for row in rows:
        company_categories.setdefault(row["companyName"], set()).add(row["category"])

    companies = {}
    departments = {}
    for company, categories in sorted(company_categories.items()):
        company_id = slug(company)
        depts = departments_for_company(company, categories)
        companies[company_id] = {
            "companyName": company,
            "companyType": "All" if len(depts) > 1 else depts[0],
            "departments": depts,
            "registrationNumber": f"ATTACHED-{slug(company).upper()}",
            "invoiceCode": company_invoice_code(company),
            "invoiceCodePrefix": company_invoice_code(company),
            "invoiceCount": 0,
            "status": "Active",
            "source": "Attached policy Excel files",
            "createdAt": now,
            "updatedAt": now,
        }
        for dept in depts:
            departments[f"{company_id}-{slug(dept)}"] = {
                "companyId": company_id,
                "companyName": company,
                "departmentName": dept,
                "status": "Active",
                "createdAt": now,
                "updatedAt": now,
            }

    products = {"policies": {}, "life_policies": {}, "general_policies": {}}
    customers = {}
    customer_policies = {}
    revenue = {}
    seen_policy_numbers = set()
    for sheet in sheet_masters:
        category = sheet["category"]
        collection = (
            "life_policies"
            if category == "Life"
            else "policies"
            if category == "Health"
            else "general_policies"
        )
        company_id = slug(sheet["companyName"])
        product_id = slug(f"{company_id}-{category}-{sheet['product']}")
        policy_code = slug(sheet["product"]).upper()[:60] or f"{category.upper()}-PORTFOLIO"
        product_doc = products[collection].setdefault(
            product_id,
            {
                "companyId": company_id,
                "companyName": sheet["companyName"],
                "planName": sheet["product"],
                "policyCode": policy_code,
                "description": f"Imported from {sheet['sourceWorkbook']} / {sheet['sourceSheet']}",
                "category": category,
                "policySection": category,
                "status": "Active",
                "renewalCommission": 0,
                "commissionRules": [],
                "searchKey": f"{sheet['product']} {policy_code} {sheet['companyName']} {category}".lower(),
                "sourceWorkbook": sheet["sourceWorkbook"],
                "sourceSheet": sheet["sourceSheet"],
                "createdAt": now,
                "updatedAt": now,
            },
        )
        if collection == "life_policies":
            product_doc.setdefault("lifeCommissions", [])
        elif collection == "policies":
            product_doc.setdefault("healthCommissions", [])
        else:
            product_doc.setdefault("generalCommissions", [{"type": "premium", "slabs": []}])

    for index, row in enumerate(rows, 1):
        company_id = slug(row["companyName"])
        category = row["category"]
        collection = (
            "life_policies"
            if category == "Life"
            else "policies"
            if category == "Health"
            else "general_policies"
        )
        product_id = slug(f"{company_id}-{category}-{row['product']}")
        policy_code = slug(row["product"]).upper()[:60] or f"{category.upper()}-PORTFOLIO"
        product_doc = products[collection].setdefault(
            product_id,
            {
                "companyId": company_id,
                "companyName": row["companyName"],
                "planName": row["product"],
                "policyCode": policy_code,
                "description": f"Imported from {row['sourceWorkbook']} / {row['sourceSheet']}",
                "category": category,
                "policySection": category,
                "status": "Active",
                "renewalCommission": 0,
                "commissionRules": [],
                "searchKey": f"{row['product']} {policy_code} {row['companyName']} {category}".lower(),
                "sourceWorkbook": row["sourceWorkbook"],
                "sourceSheet": row["sourceSheet"],
                "createdAt": now,
                "updatedAt": now,
            },
        )
        if row["commissionPercent"]:
            product_doc["commissionRules"].append(
                {
                    "label": f"Source row {row['sourceRow']}",
                    "percent": row["commissionPercent"],
                    "metric": "premium",
                }
            )
        if collection == "life_policies":
            product_doc["lifeCommissions"] = []
        elif collection == "policies":
            product_doc["healthCommissions"] = []
        else:
            product_doc["generalCommissions"] = [
                {
                    "type": "premium",
                    "slabs": [
                        {
                            "label": "Attached source commission",
                            "percent": row["commissionPercent"],
                        }
                    ]
                    if row["commissionPercent"]
                    else [],
                }
            ]

        customer_key = slug(f"{row['customerName']}-{row['phone'] or row['policyNumber']}")
        customers[customer_key] = {
            "name": row["customerName"],
            "phone": row["phone"],
            "mobile": row["phone"],
            "category": category,
            "status": "Active",
            "source": "Attached policy Excel files",
            "createdAt": now,
            "updatedAt": now,
            "searchKey": f"{row['customerName']} {row['phone']} {category}".lower(),
        }

        cp_id = slug(row["policyNumber"])
        if cp_id in seen_policy_numbers:
            cp_id = f"{cp_id}-{index}"
        seen_policy_numbers.add(cp_id)
        start = row["issueDate"]
        end = row["policyEndDate"]
        cp = {
            "customerId": customer_key,
            "customerName": row["customerName"],
            "customerMobile": row["phone"],
            "category": category,
            "companyId": company_id,
            "companyName": row["companyName"],
            "policyId": product_id,
            "policyName": row["product"],
            "productName": row["product"],
            "policyCode": policy_code,
            "policyNumber": row["policyNumber"],
            "issueDate": iso(start),
            "issueDateFormatted": fmt_date(start),
            "policyStartDate": iso(start),
            "policyEndDate": iso(end),
            "policyEndDateFormatted": fmt_date(end),
            "sumInsured": row["sumAssured"],
            "premium": row["premium"],
            "premiumAmount": row["premium"],
            "commissionPercent": row["commissionPercent"],
            "commissionAmount": row["commissionAmount"],
            "gst": row["gst"],
            "netTotal": row["netTotal"],
            "status": "Active",
            "sourceWorkbook": row["sourceWorkbook"],
            "sourceSheet": row["sourceSheet"],
            "sourceRow": row["sourceRow"],
            "createdAt": now,
            "updatedAt": now,
        }
        customer_policies[cp_id] = cp
        rev_id = cp_id
        revenue[rev_id] = {
            **cp,
            "customerPolicyId": cp_id,
            "revenue": row["commissionAmount"],
            "month": month_key(start),
            "year": start.year if start else 0,
            "settled": False,
            "settlementStatus": "open",
        }

    policies_by_customer = {}
    for policy_id, policy in customer_policies.items():
        policies_by_customer.setdefault(policy["customerId"], []).append((policy_id, policy))

    def customer_sort_key(item):
        customer_id, customer = item
        policy_dates = [
            policy.get("policyStartDate") or policy.get("issueDate") or ""
            for _, policy in policies_by_customer.get(customer_id, [])
        ]
        return (
            min([date for date in policy_dates if date] or ["9999"]),
            clean_text(customer.get("name")).lower(),
            customer_id,
        )

    lead_serial_registry = {}
    last_unique_id = ""
    for sequence, (customer_id, customer) in enumerate(
        sorted(customers.items(), key=customer_sort_key),
        1,
    ):
        lead_id, fields = lead_fields(customer.get("category") or "General", serial_year, sequence)
        last_unique_id = lead_id
        customer.update(fields)
        customer["searchKey"] = f"{customer.get('searchKey', '')} {lead_id}".strip().lower()

        for _, policy in policies_by_customer.get(customer_id, []):
            policy.update(fields)
        for row in revenue.values():
            if row.get("customerId") == customer_id:
                row.update(fields)

        lead_serial_registry[lead_id] = {
            "leadUniqueId": lead_id,
            "sequence": sequence,
            "year": serial_year,
            "category": customer.get("category") or "General",
            "categoryCode": fields["leadSerialCategoryCode"],
            "leadId": customer_id,
            "leadName": customer.get("name") or "Unnamed",
            "status": "Allocated",
            "allocatedAt": now,
            "updatedAt": now,
        }

    return {
        "insurance_companies": companies,
        "insurance_departments": departments,
        "policies": products["policies"],
        "life_policies": products["life_policies"],
        "general_policies": products["general_policies"],
        "customers": customers,
        "customer_policies": customer_policies,
        "revenue": revenue,
        "lead_serial_registry": lead_serial_registry,
        "system_counters": {
            f"lead_serial_{serial_year}": {
                "year": serial_year,
                "lastSequence": len(customers),
                "lastLeadUniqueId": last_unique_id,
                "updatedAt": now,
            }
        },
    }


def fs_value(value):
    if value is None:
        return {"nullValue": None}
    if isinstance(value, bool):
        return {"booleanValue": value}
    if isinstance(value, int):
        return {"integerValue": str(value)}
    if isinstance(value, float):
        return {"doubleValue": value}
    if isinstance(value, list):
        return {"arrayValue": {"values": [fs_value(v) for v in value]}}
    if isinstance(value, dict):
        return {"mapValue": {"fields": fs_fields(value)}}
    if isinstance(value, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z", value):
        return {"timestampValue": value}
    return {"stringValue": str(value)}


def fs_fields(data):
    return {key: fs_value(value) for key, value in data.items()}


def token():
    cfg_path = Path.home() / ".config" / "configstore" / "firebase-tools.json"
    cfg = json.loads(cfg_path.read_text())
    return cfg["tokens"]["access_token"]


def http(method, url, access_token, payload=None):
    data = None if payload is None else json.dumps(payload).encode()
    req = request.Request(
        url,
        data=data,
        method=method,
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
        },
    )
    try:
        with request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode()
            return resp.status, json.loads(body) if body else {}
    except error.HTTPError as exc:
        body = exc.read().decode()
        raise RuntimeError(f"{method} {url} failed {exc.code}: {body[:800]}")


def list_collection(collection, access_token):
    docs = []
    page_token = ""
    while True:
        url = f"{BASE_URL}/{collection}?pageSize=300"
        if page_token:
            url += f"&pageToken={page_token}"
        _, body = http("GET", url, access_token)
        docs.extend(body.get("documents", []))
        page_token = body.get("nextPageToken", "")
        if not page_token:
            return docs


def delete_collection(collection, access_token):
    docs = list_collection(collection, access_token)
    for doc in docs:
        http("DELETE", f"https://firestore.googleapis.com/v1/{doc['name']}", access_token)
    return len(docs)


def write_collection(collection, docs, access_token):
    for doc_id, data in docs.items():
        safe_id = request.pathname2url(doc_id)
        http(
            "PATCH",
            f"{BASE_URL}/{collection}/{safe_id}",
            access_token,
            {"fields": fs_fields(data)},
        )
    return len(docs)


def summarize(rows, docs, skipped, sheet_masters):
    by_file = {}
    by_sheet = {}
    by_category = {}
    by_company = {}
    missing_dates = 0
    missing_date_rows = []
    for row in rows:
        by_file[row["sourceWorkbook"]] = by_file.get(row["sourceWorkbook"], 0) + 1
        by_sheet[f"{row['sourceWorkbook']} / {row['sourceSheet']}"] = (
            by_sheet.get(f"{row['sourceWorkbook']} / {row['sourceSheet']}", 0) + 1
        )
        by_category[row["category"]] = by_category.get(row["category"], 0) + 1
        by_company[row["companyName"]] = by_company.get(row["companyName"], 0) + 1
        if not row["issueDate"] or not row["policyEndDate"]:
            missing_dates += 1
            missing_date_rows.append(
                {
                    "workbook": row["sourceWorkbook"],
                    "sheet": row["sourceSheet"],
                    "row": row["sourceRow"],
                    "policyNumber": row["policyNumber"],
                }
            )
    return {
        "parsedRows": len(rows),
        "attachedSheets": len(sheet_masters),
        "byFile": by_file,
        "bySheet": by_sheet,
        "byCategory": by_category,
        "byCompany": by_company,
        "docCounts": {k: len(v) for k, v in docs.items()},
        "missingStartOrEndDates": missing_dates,
        "missingDateRows": missing_date_rows[:20],
        "skippedSheets": skipped,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    rows, skipped, sheet_masters = parse_workbooks()
    docs = build_documents(rows, sheet_masters)
    summary = summarize(rows, docs, skipped, sheet_masters)
    print(json.dumps(summary, indent=2))

    if not args.apply:
        print("DRY RUN ONLY. Re-run with --apply to replace Firestore collections.")
        return

    access_token = token()
    print("Replacing Firestore collections...")
    deleted = {}
    for collection in COLLECTIONS_TO_REPLACE:
        deleted[collection] = delete_collection(collection, access_token)
        print(f"deleted {collection}: {deleted[collection]}")
    written = {}
    for collection in COLLECTIONS_TO_REPLACE:
        written[collection] = write_collection(collection, docs[collection], access_token)
        print(f"written {collection}: {written[collection]}")
    written["system_counters"] = write_collection(
        "system_counters",
        docs["system_counters"],
        access_token,
    )
    print(f"written system_counters: {written['system_counters']}")
    print(json.dumps({"deleted": deleted, "written": written}, indent=2))


if __name__ == "__main__":
    main()
