import argparse
import json
import math
import re
from datetime import UTC, datetime, timedelta
from pathlib import Path
from urllib import error, parse, request

from openpyxl import load_workbook


PROJECT_ID = "insurance-1178e"
API_KEY = "AIzaSyBt7a5e3CM8GwO4tNwHkvDMcw2JqAGbiQU"
BASE_URL = (
    f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}"
    "/databases/(default)/documents"
)

SOURCE_FILE = Path(
    r"C:\Users\user\Downloads\Staff Health Policies Data_Health Policies.xlsx"
)

CUSTOMER_WORKFLOW_COLLECTIONS = [
    "data_transfer_directories",
    "data_transfer_contacts",
    "data_transfer_batches",
    "telecaller_leads",
    "customers",
    "customer_policies",
    "revenue",
    "lead_serial_registry",
    "policy_serial_registry",
]

SYSTEM_COUNTER_DOCS = ["lead_serial_2026", "policy_serial_2026"]

COMPANY_ALIASES = {
    "ADITYA BIRLA SUN LIFE": "Aditya Birla Sun Life  Insurance Company Ltd",
    "CARE HEALTH": "Care Health Insurance Company Ltd",
    "MANIPAL CIGNA": "Manipal Cigna Health Insurance Co Ltd",
    "MANIPALCIGNA": "Manipal Cigna Health Insurance Co Ltd",
    "NIVA BUPA": "Niva Bupa Health Insurance Co Ltd",
    "AXIS MAX LIFE": "Axis Max Life Insurance Co Ltd",
    "MAX LIFE": "Axis Max Life Insurance Co Ltd",
    "BAJAJ LIFE": "Bajaj Life Insurance Co Ltd",
    "BAJAJ LIFE INSURANCE": "Bajaj Life Insurance Co Ltd",
    "HDFC ERGO HEALTH": "HDFC ERGO General Insurance Co Ltd(Health)",
    "ICICI LOMBARD HEALTH": "ICICI LOMBARD General Insurance Co Ltd(Health)",
    "SBI HEALTH": "SBI General Insurance Co Ltd(Health)",
    "SBI GENERAL HEALTH": "SBI General Insurance Co Ltd(Health)",
    "IFFCO TOKIO": "Iffco Tokio General Insurance Co Ltd",
    "STAR HEALTH": "Star Health & Allied Insurance Co Ltd",
    "ICICI PRUDENTIAL": "ICICI Prudential Life Insurance Co Ltd",
    "GALAXY HEALTH": "Galaxy Health Insurance Co Ltd",
    "ADITYA BIRLA HEALTH": "Aditya Birla Health  Insurance Co Ltd",
    "BAJAJ GENERAL": "Baja General Insurance Co Ltd",
    "UNITED INDIA": "United India Isurance Co Ltd",
    "NEW INDIA": "New India Assurance Co Ltd",
    "SBI LIFE": "SBI Life Insurance Co Ltd",
    "INDUS IND NIPPON": "Indus Ind Nippon Life Insurance Co Ltd",
    "HDFC ERGO MOTOR": "HDFC ERGO General Insurance Co Ltd(Motor)",
    "ICICI LOMBARD MOTOR": "ICICI LomBARD General Insurance Co Ltd(Motor)",
    "SBI MOTOR": "SBI General Insurance Co Ltd(Motor)",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\s+", " ", text)


def slug(value):
    text = clean(value).lower().replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "item"


def number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", str(value))
    if text in {"", "-", "."}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)) and not math.isnan(value):
        return datetime(1899, 12, 30) + timedelta(days=int(value))
    text = clean(value)
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def one_year_minus_one_day(start):
    try:
        return datetime(start.year + 1, start.month, start.day) - timedelta(days=1)
    except ValueError:
        return datetime(start.year + 1, start.month, 28) - timedelta(days=1)


def iso(dt):
    if not dt:
        return None
    return dt.strftime("%Y-%m-%dT00:00:00Z")


def fmt_date(dt):
    if not dt:
        return ""
    return dt.strftime("%d.%m.%Y")


def month_key(dt):
    if not dt:
        return ""
    return f"{dt.year}-{dt.month:02d}"


def category_code(category):
    key = clean(category).lower()
    return {
        "health": "H",
        "life": "L",
        "general": "G",
        "agriculture": "A",
        "agricultural": "A",
        "ecgc": "E",
    }.get(key, "G")


def lead_fields(category, year, sequence):
    lead_id = f"M{category_code(category)}{year}{sequence:05d}"
    return {
        "leadUniqueId": lead_id,
        "uniqueLeadId": lead_id,
        "leadSerialNumber": lead_id,
        "leadSerialSequence": sequence,
        "leadSerialYear": year,
        "leadSerialCategoryCode": category_code(category),
    }


def policy_serial_fields(category, year, sequence):
    serial = f"M{category_code(category)}{str(year % 100).zfill(2)}{sequence:06d}"
    return {
        "serialNumber": serial,
        "policySerialNumber": serial,
        "serialSequence": sequence,
        "serialYear": year,
        "serialCategoryCode": category_code(category),
    }


def normalize_company(name, category, code_companies):
    raw = clean(name)
    key = raw.upper()
    context_key = f"{key} {category.upper()}".strip()
    for alias_key in (context_key, key):
        if alias_key in COMPANY_ALIASES:
            return COMPANY_ALIASES[alias_key]
    for known in code_companies:
        if key and key in known.upper():
            return known
    return raw or "Unknown Insurance Company"


def company_type(category):
    if category == "Health":
        return "Health"
    if category == "Life":
        return "Life"
    return "General"


def product_collection(category):
    if category == "Life":
        return "life_policies"
    if category == "Health":
        return "policies"
    return "general_policies"


def find_header(ws):
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        values = [
            clean(ws.cell(row_idx, col).value).lower()
            for col in range(1, ws.max_column + 1)
        ]
        joined = " ".join(values)
        if "sl no" in joined and "policy" in joined:
            return row_idx, values
    return None, []


def header_map(headers):
    result = {}
    for idx, label in enumerate(headers):
        if "sl no" in label:
            result["sl_no"] = idx
        elif "contact no" in label or "mobile" in label or "phone" in label:
            result["phone"] = idx
        elif "policy no" in label:
            result["policy_no"] = idx
        elif "insurer name" in label or "insurance co" in label or "company name" in label:
            result["company"] = idx
        elif "name" in label and ("person" in label or "company" not in label):
            result["name"] = idx
        elif "issued" in label:
            result["issue_date"] = idx
        elif "policy end" in label or "end date" in label or "expiry" in label:
            result["end_date"] = idx
        elif "sum insured" in label or "sum assured" in label:
            result["sum"] = idx
        elif "premium" in label:
            result["premium"] = idx
        elif "remarks" in label or "contact person" in label:
            result["notes"] = idx
        elif "commission" in label and "%" in label:
            result["commission_percent"] = idx
        elif label == "commission":
            result["commission_percent"] = idx
        elif "total commission" in label or label == "total":
            result["commission_amount"] = idx
        elif label == "received":
            result["received"] = idx
        elif "received date" in label:
            result["received_date"] = idx
    return result


def parse_code_sheet(wb):
    if "Insurance Codes" not in wb.sheetnames:
        return {}, {}
    ws = wb["Insurance Codes"]
    header_idx, headers = find_header(ws)
    if not header_idx:
        header_idx = 2
        headers = [clean(ws.cell(header_idx, c).value).lower() for c in range(1, ws.max_column + 1)]
    hm = header_map(headers)
    companies = {}
    by_name = {}
    for r in range(header_idx + 1, ws.max_row + 1):
        name = clean(ws.cell(r, 2).value)
        if not name:
            continue
        code = clean(ws.cell(r, 3).value)
        contact = clean(ws.cell(r, 4).value)
        mobile = clean(ws.cell(r, 5).value)
        category = "General"
        text = name.lower()
        if "life" in text:
            category = "Life"
        elif "health" in text:
            category = "Health"
        elif "agricultur" in text:
            category = "Agriculture"
        elif "export credit" in text or "ecgc" in text:
            category = "ECGC"
        company_id = slug(name)
        doc = {
            "companyName": name,
            "companyType": company_type(category) if category != "Agriculture" else "Agriculture",
            "departments": [category],
            "registrationNumber": code,
            "invoiceCode": f"MAKK_{slug(name).upper()[:12]}",
            "invoiceCodePrefix": f"MAKK_{slug(name).upper()[:12]}",
            "invoiceCount": 0,
            "contactPerson": contact,
            "contactMobile": mobile,
            "status": "Active",
            "source": SOURCE_FILE.name,
            "updatedAt": datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ"),
        }
        companies[company_id] = doc
        by_name[name] = doc
    return companies, by_name


def parse_policy_rows(wb, code_company_names):
    configs = {
        "Health Policies": ("Health", "Health Policies"),
        "Motor Policies": ("General", "Motor Policies"),
        "Life Insurance Policies": ("Life", "Life Insurance Policies"),
    }
    rows = []
    skipped = []
    for sheet_name, (category, product_name) in configs.items():
        if sheet_name not in wb.sheetnames:
            skipped.append({"sheet": sheet_name, "reason": "sheet missing"})
            continue
        ws = wb[sheet_name]
        header_idx, headers = find_header(ws)
        if not header_idx:
            skipped.append({"sheet": sheet_name, "reason": "header not found"})
            continue
        hm = header_map(headers)
        required = {"sl_no", "name", "phone", "policy_no", "issue_date", "end_date", "company"}
        missing = sorted(required - set(hm))
        if missing:
            skipped.append({"sheet": sheet_name, "reason": f"missing {missing}"})
            continue
        for r in range(header_idx + 1, ws.max_row + 1):
            values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            sl_no = clean(values[hm["sl_no"]])
            if not re.fullmatch(r"\d+", sl_no):
                continue
            name = clean(values[hm["name"]])
            phone = clean(values[hm["phone"]])
            policy_no = clean(values[hm["policy_no"]]).upper()
            if not name or not phone or not policy_no:
                skipped.append({"sheet": sheet_name, "row": r, "reason": "missing mandatory name/phone/policy"})
                continue
            issue = parse_date(values[hm["issue_date"]])
            end = parse_date(values[hm["end_date"]])
            corrected_end = False
            if issue and not end:
                end = one_year_minus_one_day(issue)
                corrected_end = True
            if end and not issue:
                issue = datetime(end.year - 1, end.month, min(end.day + 1, 28))
            if issue and end and end <= issue:
                end = one_year_minus_one_day(issue)
                corrected_end = True
            company = normalize_company(values[hm["company"]], category, code_company_names)
            percent = number(values[hm["commission_percent"]]) if "commission_percent" in hm else 0.0
            if 0 < percent <= 1:
                percent *= 100
            premium = number(values[hm["premium"]]) if "premium" in hm else 0.0
            commission_amount = number(values[hm["commission_amount"]]) if "commission_amount" in hm else 0.0
            if not commission_amount and percent:
                commission_amount = round((premium * percent) / 100, 2)
            rows.append(
                {
                    "sourceSheet": sheet_name,
                    "sourceRow": r,
                    "sourceSlNo": int(sl_no),
                    "category": category,
                    "productName": product_name,
                    "customerName": name,
                    "phone": phone,
                    "policyNumber": policy_no,
                    "issueDate": issue,
                    "policyEndDate": end,
                    "dateCorrected": corrected_end,
                    "sumInsured": clean(values[hm["sum"]]) if "sum" in hm else "",
                    "sumInsuredAmount": number(values[hm["sum"]]) if "sum" in hm else 0.0,
                    "premium": premium,
                    "companyName": company,
                    "rawCompanyName": clean(values[hm["company"]]),
                    "notes": clean(values[hm["notes"]]) if "notes" in hm else "",
                    "commissionPercent": percent,
                    "commissionAmount": commission_amount,
                    "received": clean(values[hm["received"]]) if "received" in hm else "",
                    "receivedDate": parse_date(values[hm["received_date"]]) if "received_date" in hm else None,
                }
            )
    return rows, skipped


def build_documents(rows, code_companies):
    now = datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    serial_year = 2026

    companies = dict(code_companies)
    departments = {}
    products = {"policies": {}, "life_policies": {}, "general_policies": {}}
    customers = {}
    customer_policies = {}
    revenue = {}
    lead_registry = {}
    policy_registry = {}

    for row in rows:
        company_id = slug(row["companyName"])
        category = row["category"]
        existing = companies.get(company_id, {})
        departments_set = set(existing.get("departments", []))
        departments_set.add(category)
        companies[company_id] = {
            **existing,
            "companyName": row["companyName"],
            "companyType": "All" if len(departments_set) > 1 else company_type(category),
            "departments": sorted(departments_set),
            "registrationNumber": existing.get("registrationNumber", ""),
            "invoiceCode": existing.get("invoiceCode") or f"MAKK_{slug(row['companyName']).upper()[:12]}",
            "invoiceCodePrefix": existing.get("invoiceCodePrefix") or f"MAKK_{slug(row['companyName']).upper()[:12]}",
            "invoiceCount": existing.get("invoiceCount", 0),
            "status": "Active",
            "source": SOURCE_FILE.name,
            "updatedAt": now,
        }
        departments[f"{company_id}-{slug(category)}"] = {
            "companyId": company_id,
            "companyName": row["companyName"],
            "departmentName": category,
            "status": "Active",
            "source": SOURCE_FILE.name,
            "updatedAt": now,
        }
        collection = product_collection(category)
        product_id = slug(f"{company_id}-{category}-{row['productName']}")
        policy_code = f"{category_code(category)}-{slug(row['productName']).upper()}"
        products[collection].setdefault(
            product_id,
            {
                "companyId": company_id,
                "companyName": row["companyName"],
                "planName": f"{row['companyName']} {row['productName']}",
                "policyCode": policy_code,
                "description": f"Imported from {SOURCE_FILE.name} / {row['sourceSheet']}",
                "category": category,
                "policySection": category,
                "status": "Active",
                "renewalCommission": 0,
                "commissionRules": [],
                "healthCommissions": [] if category == "Health" else None,
                "lifeCommissions": [] if category == "Life" else None,
                "generalCommissions": [{"type": "premium", "slabs": []}] if category == "General" else None,
                "searchKey": f"{row['companyName']} {row['productName']} {policy_code} {category}".lower(),
                "sourceWorkbook": SOURCE_FILE.name,
                "sourceSheet": row["sourceSheet"],
                "createdAt": now,
                "updatedAt": now,
            },
        )

    seen_policy_ids = set()
    for row in rows:
        customer_id = slug(f"{row['customerName']}-{row['phone']}")
        notes = row["notes"]
        customer = customers.setdefault(
            customer_id,
            {
                "fullName": row["customerName"],
                "name": row["customerName"],
                "mobileNumber": row["phone"],
                "phone": row["phone"],
                "mobile": row["phone"],
                "customerType": "Individual",
                "customerCategory": row["category"],
                "category": row["category"],
                "leadStatus": "Green",
                "status": "Active",
                "policyLinkedManually": False,
                "source": SOURCE_FILE.name,
                "createdAt": now,
                "updatedAt": now,
                "searchKey": f"{row['customerName']} {row['phone']} {row['category']}".lower(),
                "notes": "",
                "importedPolicyNotes": [],
            },
        )
        if notes:
            customer["importedPolicyNotes"].append(
                {
                    "policyNumber": row["policyNumber"],
                    "policyName": row["productName"],
                    "companyName": row["companyName"],
                    "notes": notes,
                }
            )
            customer["notes"] = "\n".join(
                f"{item['policyNumber']} - {item['notes']}"
                for item in customer["importedPolicyNotes"]
            )

        policy_id = slug(row["policyNumber"])
        if policy_id in seen_policy_ids:
            policy_id = f"{policy_id}-{row['sourceRow']}"
        seen_policy_ids.add(policy_id)
        collection = product_collection(row["category"])
        company_id = slug(row["companyName"])
        product_id = slug(f"{company_id}-{row['category']}-{row['productName']}")
        policy_name = f"{row['companyName']} {row['productName']}"
        policy_code = f"{category_code(row['category'])}-{slug(row['productName']).upper()}"
        policy_fields = {
            "customerId": customer_id,
            "customerName": row["customerName"],
            "customerMobile": row["phone"],
            "category": row["category"],
            "customerCategory": row["category"],
            "companyId": company_id,
            "companyName": row["companyName"],
            "rawCompanyName": row["rawCompanyName"],
            "policyId": product_id,
            "policyCollection": collection,
            "policyName": policy_name,
            "productName": row["productName"],
            "policyCode": policy_code,
            "policyNumber": row["policyNumber"],
            "issueDate": iso(row["issueDate"]),
            "issueDateFormatted": fmt_date(row["issueDate"]),
            "policyStartDate": iso(row["issueDate"]),
            "policyEndDate": iso(row["policyEndDate"]),
            "policyEndDateFormatted": fmt_date(row["policyEndDate"]),
            "dateCorrectedOnImport": row["dateCorrected"],
            "sumInsured": row["sumInsured"],
            "sumAssured": row["sumInsuredAmount"],
            "premium": row["premium"],
            "premiumAmount": row["premium"],
            "commissionPercent": row["commissionPercent"],
            "commissionAmount": row["commissionAmount"],
            "totalCommission": row["commissionAmount"],
            "received": row["received"],
            "receivedDate": iso(row["receivedDate"]),
            "receivedDateFormatted": fmt_date(row["receivedDate"]),
            "status": "Active",
            "notes": row["notes"],
            "renewalNotes": row["notes"],
            "sourceWorkbook": SOURCE_FILE.name,
            "sourceSheet": row["sourceSheet"],
            "sourceRow": row["sourceRow"],
            "sourceSlNo": row["sourceSlNo"],
            "createdAt": now,
            "updatedAt": now,
        }
        customer_policies[policy_id] = policy_fields
        revenue[policy_id] = {
            **policy_fields,
            "customerPolicyId": policy_id,
            "month": month_key(row["issueDate"]),
            "year": row["issueDate"].year if row["issueDate"] else serial_year,
            "revenue": row["commissionAmount"],
            "settled": False,
            "settlementStatus": "open",
        }

    policies_by_customer = {}
    for policy_id, policy in customer_policies.items():
        policies_by_customer.setdefault(policy["customerId"], []).append(policy)

    def customer_sort(item):
        customer_id, customer = item
        dates = [
            p.get("policyStartDate") or p.get("issueDate") or ""
            for p in policies_by_customer.get(customer_id, [])
        ]
        return (min([d for d in dates if d] or ["9999"]), clean(customer["fullName"]).lower(), customer_id)

    last_lead_id = ""
    for sequence, (customer_id, customer) in enumerate(sorted(customers.items(), key=customer_sort), 1):
        first_policy = sorted(
            policies_by_customer.get(customer_id, []),
            key=lambda p: p.get("policyStartDate") or "9999",
        )[0]
        fields = lead_fields(first_policy["category"], serial_year, sequence)
        last_lead_id = fields["leadUniqueId"]
        customer.update(fields)
        customer["customerCategory"] = first_policy["category"]
        customer["category"] = first_policy["category"]
        customer["searchKey"] = f"{customer['searchKey']} {last_lead_id}".lower()
        for policy in policies_by_customer.get(customer_id, []):
            policy.update(fields)
            revenue[slug(policy["policyNumber"]) if slug(policy["policyNumber"]) in revenue else next(
                key for key, value in revenue.items() if value["policyNumber"] == policy["policyNumber"]
            )].update(fields)
        lead_registry[last_lead_id] = {
            "leadUniqueId": last_lead_id,
            "sequence": sequence,
            "year": serial_year,
            "category": first_policy["category"],
            "categoryCode": fields["leadSerialCategoryCode"],
            "leadId": customer_id,
            "leadName": customer["fullName"],
            "status": "Allocated",
            "allocatedAt": now,
            "updatedAt": now,
        }

    def policy_sort(item):
        policy_id, policy = item
        return (
            policy.get("policyStartDate") or policy.get("issueDate") or "9999",
            clean(policy.get("customerName")).lower(),
            policy.get("policyNumber") or policy_id,
        )

    last_policy_serial = ""
    for sequence, (policy_id, policy) in enumerate(
        sorted(customer_policies.items(), key=policy_sort),
        1,
    ):
        serial = policy_serial_fields(policy["category"], serial_year, sequence)
        last_policy_serial = serial["serialNumber"]
        policy.update(serial)
        revenue[policy_id].update(serial)
        policy_registry[last_policy_serial] = {
            "serialNumber": last_policy_serial,
            "sequence": sequence,
            "year": serial_year,
            "category": policy["category"],
            "categoryCode": serial["serialCategoryCode"],
            "policyId": policy_id,
            "customerId": policy["customerId"],
            "customerName": policy["customerName"],
            "status": "Allocated",
            "allocatedAt": now,
        }

    cleaned_products = {}
    for collection, docs in products.items():
        cleaned_products[collection] = {
            doc_id: {k: v for k, v in data.items() if v is not None}
            for doc_id, data in docs.items()
        }

    return {
        "insurance_companies": companies,
        "insurance_departments": departments,
        "policies": cleaned_products["policies"],
        "life_policies": cleaned_products["life_policies"],
        "general_policies": cleaned_products["general_policies"],
        "customers": customers,
        "customer_policies": customer_policies,
        "revenue": revenue,
        "lead_serial_registry": lead_registry,
        "policy_serial_registry": policy_registry,
        "system_counters": {
            f"lead_serial_{serial_year}": {
                "year": serial_year,
                "lastSequence": len(customers),
                "lastLeadUniqueId": last_lead_id,
                "updatedAt": now,
            },
            f"policy_serial_{serial_year}": {
                "year": serial_year,
                "lastSequence": len(customer_policies),
                "lastSerialNumber": last_policy_serial,
                "updatedAt": now,
            },
        },
    }


def fs_value(value):
    if value is None:
        return {"nullValue": None}
    if isinstance(value, bool):
        return {"booleanValue": value}
    if isinstance(value, int):
        return {"integerValue": str(value)}
    if isinstance(value, float):
        return {"doubleValue": value}
    if isinstance(value, list):
        return {"arrayValue": {"values": [fs_value(v) for v in value]}}
    if isinstance(value, dict):
        return {"mapValue": {"fields": fs_fields(value)}}
    if isinstance(value, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z", value):
        return {"timestampValue": value}
    return {"stringValue": str(value)}


def fs_fields(data):
    return {key: fs_value(value) for key, value in data.items()}


def token():
    csv_path = Path(__file__).with_name("employee_credentials_2026-07-17.csv")
    rows = [
        [part.strip() for part in line.split(",")]
        for line in csv_path.read_text().splitlines()[1:]
        if line.strip()
    ]
    operator = next((row for row in rows if row and row[0] == "manager"), rows[0] if rows else None)
    if not operator or len(operator) < 4:
        raise RuntimeError("No manager credential is available for authenticated import.")
    email, password = operator[2], operator[3]
    payload = json.dumps(
        {"email": email, "password": password, "returnSecureToken": True}
    ).encode()
    req = request.Request(
        f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={API_KEY}",
        data=payload,
        method="POST",
        headers={"Content-Type": "application/json"},
    )
    try:
        with request.urlopen(req, timeout=60) as resp:
            body = json.loads(resp.read().decode())
    except error.HTTPError as exc:
        body = exc.read().decode()
        raise RuntimeError(f"Manager Firebase sign-in failed: {body[:800]}")
    return body["idToken"]


def http(method, url, access_token, payload=None):
    data = None if payload is None else json.dumps(payload).encode()
    req = request.Request(
        url,
        data=data,
        method=method,
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
        },
    )
    try:
        with request.urlopen(req, timeout=90) as resp:
            body = resp.read().decode()
            return resp.status, json.loads(body) if body else {}
    except error.HTTPError as exc:
        body = exc.read().decode()
        raise RuntimeError(f"{method} {url} failed {exc.code}: {body[:1200]}")


def list_collection(collection, access_token):
    docs = []
    page_token = ""
    while True:
        url = f"{BASE_URL}/{collection}?pageSize=500"
        if page_token:
            url += f"&pageToken={page_token}"
        status, body = http("GET", url, access_token)
        docs.extend(body.get("documents", []))
        page_token = body.get("nextPageToken", "")
        if not page_token:
            return docs


def delete_doc_path(path, access_token):
    try:
        http("DELETE", f"{BASE_URL}/{path}", access_token)
        return True
    except RuntimeError as exc:
        if "failed 404" in str(exc):
            return False
        raise


def delete_collection(collection, access_token):
    docs = list_collection(collection, access_token)
    for doc in docs:
        http("DELETE", f"https://firestore.googleapis.com/v1/{doc['name']}", access_token)
    return len(docs)


def write_collection(collection, docs, access_token, merge=False):
    for doc_id, data in docs.items():
        safe_id = parse.quote(doc_id, safe="")
        url = f"{BASE_URL}/{collection}/{safe_id}"
        if merge:
            for field in data:
                url += ("&" if "?" in url else "?") + f"updateMask.fieldPaths={parse.quote(field)}"
        http("PATCH", url, access_token, {"fields": fs_fields(data)})
    return len(docs)


def summarize(rows, docs, skipped):
    by_sheet = {}
    by_category = {}
    by_company = {}
    corrected_dates = 0
    missing_dates = []
    for row in rows:
        by_sheet[row["sourceSheet"]] = by_sheet.get(row["sourceSheet"], 0) + 1
        by_category[row["category"]] = by_category.get(row["category"], 0) + 1
        by_company[row["companyName"]] = by_company.get(row["companyName"], 0) + 1
        if row["dateCorrected"]:
            corrected_dates += 1
        if not row["issueDate"] or not row["policyEndDate"]:
            missing_dates.append(
                {
                    "sheet": row["sourceSheet"],
                    "row": row["sourceRow"],
                    "policyNumber": row["policyNumber"],
                }
            )
    return {
        "source": str(SOURCE_FILE),
        "parsedPolicyRows": len(rows),
        "uniqueCustomers": len(docs["customers"]),
        "linkedCustomerPolicies": len(docs["customer_policies"]),
        "bySheet": by_sheet,
        "byCategory": by_category,
        "byCompany": by_company,
        "correctedEndDates": corrected_dates,
        "missingStartOrEndDates": missing_dates,
        "writeCounts": {key: len(value) for key, value in docs.items() if isinstance(value, dict)},
        "skipped": skipped[:50],
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    wb = load_workbook(SOURCE_FILE, data_only=True)
    code_companies, code_by_name = parse_code_sheet(wb)
    rows, skipped = parse_policy_rows(wb, code_by_name.keys())
    docs = build_documents(rows, code_companies)
    summary = summarize(rows, docs, skipped)
    print(json.dumps(summary, indent=2))
    if not args.apply:
        print("DRY RUN ONLY. Re-run with --apply to refresh Firebase.")
        return

    access_token = token()
    deleted = {}
    for collection in CUSTOMER_WORKFLOW_COLLECTIONS:
        deleted[collection] = delete_collection(collection, access_token)
        print(f"deleted {collection}: {deleted[collection]}")
    for doc_id in SYSTEM_COUNTER_DOCS:
        deleted[f"system_counters/{doc_id}"] = delete_doc_path(f"system_counters/{doc_id}", access_token)
        print(f"deleted system_counters/{doc_id}: {deleted[f'system_counters/{doc_id}']}")

    written = {}
    for collection in ("insurance_companies", "insurance_departments", "policies", "life_policies", "general_policies"):
        written[collection] = write_collection(collection, docs[collection], access_token, merge=True)
        print(f"upserted {collection}: {written[collection]}")
    for collection in ("customers", "customer_policies", "revenue", "lead_serial_registry", "policy_serial_registry", "system_counters"):
        written[collection] = write_collection(collection, docs[collection], access_token)
        print(f"written {collection}: {written[collection]}")
    print(json.dumps({"deleted": deleted, "written": written}, indent=2))


if __name__ == "__main__":
    main()
